# 导入模块
import tkinter as tk
from tkinter import messagebox
from tkinter.ttk import *
import openpyxl

'''
下面是生成查看学生信息页面的代码
'''
def ModifyRoot():
    modifyRoot = tk.Tk()
    modifyRoot.title('欢迎来到学生管理系统！')
    modifyRoot.geometry('600x350+300+300')
    modifyRoot['bg'] = 'lightblue'
    modifyRoot.attributes('-alpha', 0.9)

    # 点击"提交"按钮的操作
    def clickSubR():
        # 首先，打开Excel工作簿
        data = openpyxl.load_workbook('../lib/Students.xlsx')
        # 获取第一个工作表
        table = data.active

        # 获取输入的学号和成绩
        student_id = int(entryID.get())
        score = int(entryScore.get())
    
        # 要更新的列索引，假设学号在第一列，成绩分别在第五、第六和第七列
        column_index = None
        if comb.get() == '高数':
            column_index = 5
        elif comb.get() == 'C':
            column_index = 6
        elif comb.get() == 'Python':
            column_index = 7

        # 查找学号对应的行
        for row in table.iter_rows(min_row=2, max_col=1, max_row=table.max_row):
            cell = row[0]  # 获取学号所在的单元格
            if cell.value == student_id:
                # 如果找到对应的学号，更新成绩
                table.cell(row=cell.row, column=column_index).value = score
                break
        else:
            # 如果没有找到对应的学号，显示错误信息
            messagebox.showerror('错误', '未找到学号对应的学生记录！')
            data.close()
            return

        # 保存修改
        data.save('../lib/Students.xlsx')
        # 关闭工作簿
        data.close()
        # 显示成功信息
        messagebox.showinfo('提示', '成绩更新成功！')

    # 点击"返回"按钮的操作
    def clickReturnR():
        modifyRoot.destroy()

    # 初始化控件
    idRoot = tk.Frame(modifyRoot, bg='lightblue')
    idRoot.pack()
    subjectRoot = tk.Frame(modifyRoot, bg='lightblue')
    subjectRoot.pack()
    scoreRoot = tk.Frame(modifyRoot, bg='lightblue')
    scoreRoot.pack()
    btnRoot = tk.Frame(modifyRoot, bg='lightblue')
    btnRoot.pack()

    # 输入获取
    var_id = tk.StringVar()
    var_subject = tk.StringVar()
    var_score = tk.StringVar()

    # 学号控件
    tk.Label(idRoot, text='学号', bg='lightblue').grid(row=1, column=0, ipady=30, ipadx=10)
    entryID = Entry(idRoot, textvariable=var_id)
    entryID.grid(row=1, column=1, columnspan=3)

    # 科目控件
    tk.Label(subjectRoot, text='科目', bg='lightblue').grid(row=2, column=0, ipady=30, ipadx=10)
    comb = Combobox(subjectRoot, textvariable=var_subject, values=['高数', 'C', 'Python'], width=18)
    comb.grid(row=2, column=1)

    # 成绩控件
    tk.Label(scoreRoot, text='成绩', bg='lightblue').grid(row=3, column=0, ipady=30, ipadx=10)
    entryScore = Entry(scoreRoot, textvariable=var_score)
    entryScore.grid(row=3, column=1, columnspan=3)

    # 按钮显示
    Button(btnRoot, text='提交', width=6, command=clickSubR).grid(row=4, column=0)
    Button(btnRoot, text='返回', width=6, command=clickReturnR).grid(row=4, column=1)

    modifyRoot.mainloop()

if __name__ == '__main__':
    ModifyRoot()